#!/usr/bin/env python3
"""
MEBODY V1 — Excel 스펙 vs 앱 스냅샷 (및 선택적 Supabase DB) 검증

Usage:
  python3 scripts/verify-v1-excel-spec.py
  EXCEL_PATH=/path/to.xlsx python3 scripts/verify-v1-excel-spec.py
  VERIFY_DB=1 python3 scripts/verify-v1-excel-spec.py   # needs .env.local + openpyxl + supabase via curl optional

Requires: openpyxl (pip install openpyxl)
"""

from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = Path.home() / "Downloads" / "MEBODY_V1_2차문항_최종_개발명세.xlsx"

AXIS_MAP = {
    "1": "neck",
    "2": "shoulder",
    "3": "pelvis",
    "4": "flexibility",
}


def die(msg: str, code: int = 1) -> None:
    print(f"FAIL: {msg}")
    sys.exit(code)


def load_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return __import__("openpyxl")
    except ImportError:
        die("openpyxl required. Run: python3 -m venv .venv && .venv/bin/pip install openpyxl")


def strip_trailing_commas(text: str) -> str:
    return re.sub(r",(\s*[}\]])", r"\1", text)


def parse_ts_array(path: Path, export_name: str) -> list:
    text = path.read_text(encoding="utf-8")
    marker = f"export const {export_name} = ["
    start = text.index(marker)
    start = text.index("[", start)
    end = text.index("\n]", start)
    raw = strip_trailing_commas(text[start : end + 2])
    return json.loads(raw)


def map_excel_axis(raw) -> str | None:
    if raw is None or raw == "":
        return None
    s = str(raw)
    for digit, key in AXIS_MAP.items():
        if digit in s:
            return key
    return None


def load_excel_mapping(wb) -> dict:
    ws = wb["⑪ 점수 Mapping DB"]
    rows = {}
    for r in range(3, ws.max_row + 1):
        qid = ws.cell(r, 2).value
        choice = ws.cell(r, 3).value
        if not qid or not choice:
            continue
        key = f"{qid}_{choice}"
        rows[key] = {
            "question_code": str(qid),
            "choice": str(choice),
            "choice_summary": ws.cell(r, 4).value,
            "axis": map_excel_axis(ws.cell(r, 5).value),
            "direction": ws.cell(r, 6).value or None,
            "axis_weight": int(ws.cell(r, 7).value or 0),
            "axis_anchor": ws.cell(r, 8).value,
            "axis_priority": ws.cell(r, 9).value,
            "score_recovery": int(ws.cell(r, 10).value or 0),
            "score_strength": int(ws.cell(r, 11).value or 0),
            "score_mobility": int(ws.cell(r, 12).value or 0),
            "score_balance": int(ws.cell(r, 13).value or 0),
            "aux_tag": ws.cell(r, 15).value,
        }
    return rows


def load_excel_master(wb) -> dict:
    ws = wb["⑩ 문항 Master"]
    rows = {}
    for r in range(3, ws.max_row + 1):
        qid = ws.cell(r, 2).value
        if not qid:
            continue
        rows[str(qid)] = {
            "question_code": str(qid),
            "sort_order": int(ws.cell(r, 1).value or 0),
            "part": ws.cell(r, 3).value,
            "title": ws.cell(r, 4).value or "",
            "axis_anchor": ws.cell(r, 7).value,
            "axis_priority": ws.cell(r, 8).value,
            "question_text": ws.cell(r, 11).value or "",
            "instruction": ws.cell(r, 12).value or "",
        }
    return rows


def load_env_local() -> dict:
    path = ROOT / ".env.local"
    if not path.exists():
        return {}
    env = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k] = v
    return env


def fetch_supabase(table: str, select: str, filters: list[tuple[str, str]]) -> list | None:
    """Minimal REST fetch via urllib (no supabase-py)."""
    import urllib.parse
    import urllib.request

    env = load_env_local()
    url = env.get("VITE_SUPABASE_URL")
    key = env.get("VITE_SUPABASE_ANON_KEY")
    if not url or not key:
        return None

    qs = [("select", select)]
    for col, val in filters:
        qs.append((col, f"eq.{val}"))
    query = urllib.parse.urlencode(qs)
    req = urllib.request.Request(
        f"{url.rstrip('/')}/rest/v1/{table}?{query}",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(req, timeout=20) as resp:
        return json.loads(resp.read().decode("utf-8"))


def main() -> None:
    openpyxl = load_openpyxl()
    excel_path = Path(os.environ.get("EXCEL_PATH", str(DEFAULT_EXCEL)))
    if not excel_path.exists():
        die(f"Excel not found: {excel_path}")

    print(f"excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    excel_map = load_excel_mapping(wb)
    excel_q = load_excel_master(wb)
    app_map_list = parse_ts_array(ROOT / "src/data/v1ScoreMapping.ts", "V1_CHOICE_SCORES")
    app_q_list = parse_ts_array(ROOT / "src/data/v1QuestionsSnapshot.ts", "V1_QUESTIONS_SNAPSHOT")
    app_map = {f"{r['question_code']}_{r['choice']}": r for r in app_map_list}
    app_q = {r["question_code"]: r for r in app_q_list}

    errors: list[str] = []

    if len(excel_map) != 96:
        errors.append(f"excel mapping rows={len(excel_map)} expected 96")
    if len(app_map) != 96:
        errors.append(f"app mapping rows={len(app_map)} expected 96")
    if len(excel_q) != 32:
        errors.append(f"excel master questions={len(excel_q)} expected 32")
    if len(app_q) != 32:
        errors.append(f"app questions={len(app_q)} expected 32")

    if set(excel_map) != set(app_map):
        errors.append(f"mapping keys differ excel_only={sorted(set(excel_map)-set(app_map))[:5]} app_only={sorted(set(app_map)-set(excel_map))[:5]}")

    score_fields = (
        "axis_weight",
        "score_recovery",
        "score_strength",
        "score_mobility",
        "score_balance",
    )
    score_diffs = 0
    for key in sorted(set(excel_map) & set(app_map)):
        e, a = excel_map[key], app_map[key]
        for f in score_fields:
            if int(e.get(f) or 0) != int(a.get(f) or 0):
                score_diffs += 1
                if score_diffs <= 10:
                    errors.append(f"score {key}.{f}: excel={e.get(f)} app={a.get(f)}")
        ed = e.get("direction") or None
        ad = a.get("direction") or None
        if ed != ad and (e.get("axis_weight") or 0) > 0:
            score_diffs += 1
            errors.append(f"direction {key}: excel={ed} app={ad}")
        if (e.get("axis_weight") or 0) > 0 and e.get("axis") != a.get("axis"):
            score_diffs += 1
            errors.append(f"axis {key}: excel={e.get('axis')} app={a.get('axis')}")

    if set(excel_q) != set(app_q):
        errors.append(
            f"question codes differ excel_only={sorted(set(excel_q)-set(app_q))} "
            f"app_only={sorted(set(app_q)-set(excel_q))}"
        )

    q_diffs = 0
    for code in sorted(set(excel_q) & set(app_q)):
        e, a = excel_q[code], app_q[code]
        for ef, af in (
            ("question_text", "question_text"),
            ("title", "title"),
            ("instruction", "instruction"),
        ):
            if (e.get(ef) or "") != (a.get(af) or ""):
                q_diffs += 1
                if q_diffs <= 10:
                    errors.append(f"question {code}.{ef} mismatch")

    print(f"mapping: excel={len(excel_map)} app={len(app_map)} score_diffs={score_diffs}")
    print(f"questions: excel={len(excel_q)} app={len(app_q)} field_diffs={q_diffs}")

    db_checked = False
    # Optional DB compare
    if os.environ.get("VERIFY_DB") == "1":
        env = load_env_local()
        if not env.get("VITE_SUPABASE_URL") or not env.get("VITE_SUPABASE_ANON_KEY"):
            print("SKIP DB: .env.local missing URL/anon key")
        else:
            try:
                db_q = fetch_supabase(
                    "questions",
                    "question_code,question_text,title,instruction,sort_order",
                    [("question_set", "mebody_v1_32"), ("is_active", "true")],
                )
                db_m = fetch_supabase(
                    "question_choice_scores",
                    "question_code,choice,axis_weight,score_recovery,score_strength,score_mobility,score_balance",
                    [("question_set", "mebody_v1_32")],
                )
            except Exception as exc:  # noqa: BLE001
                print(f"SKIP DB: network/auth error ({exc})")
                db_q = None
                db_m = None

            if db_q is not None:
                db_checked = True
                print(f"db questions={len(db_q)} choice_scores={len(db_m or [])}")
                if len(db_q) != 32:
                    errors.append(f"db questions={len(db_q)} expected 32")
                if db_m is not None and len(db_m) != 96:
                    errors.append(f"db choice_scores={len(db_m)} expected 96")
                db_map = {f"{r['question_code']}_{r['choice']}": r for r in (db_m or [])}
                for key in sorted(set(app_map) & set(db_map)):
                    a, d = app_map[key], db_map[key]
                    for f in score_fields:
                        if int(a.get(f) or 0) != int(d.get(f) or 0):
                            errors.append(f"db score {key}.{f}: app={a.get(f)} db={d.get(f)}")

    if errors:
        print(f"\n{len(errors)} issue(s):")
        for e in errors[:40]:
            print(" -", e)
        die("Excel/app(/DB) verification failed")

    print("PASS: Excel ⑩/⑪ matches app snapshots" + (" and DB" if db_checked else ""))


if __name__ == "__main__":
    main()
